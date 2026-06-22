"""
portfolio_review.py  —  Valumonk Client Portfolio Review Generator
Generates both PDF (landscape) and Excel (multi-sheet) reports.
Returns a zip file containing both.
"""

import os, re, io, zipfile
import pandas as pd
import numpy as np
from datetime import datetime
from difflib import SequenceMatcher
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable, KeepTogether
)
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics.charts.piecharts import Pie
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Brand colours ─────────────────────────────────────────────
NAVY   = colors.HexColor('#1E2B4A')
GREEN  = colors.HexColor('#2DB84B')
LGREEN = colors.HexColor('#d1fae5')
LGREY  = colors.HexColor('#f1f5f9')
MGREY  = colors.HexColor('#64748b')
WHITE  = colors.white
RED    = colors.HexColor('#dc2626')
AMBER  = colors.HexColor('#d97706')
DGREEN = colors.HexColor('#16a34a')

PIE_COLS = [
    colors.HexColor('#1E2B4A'), colors.HexColor('#2DB84B'),
    colors.HexColor('#2563eb'), colors.HexColor('#d97706'),
    colors.HexColor('#dc2626'), colors.HexColor('#7c3aed'),
    colors.HexColor('#0891b2'), colors.HexColor('#be185d'),
    colors.HexColor('#065f46'), colors.HexColor('#92400e'),
]

COLS = {
    'name':    ['Scheme Name', 'Scheme NAV Name', 'Fund'],
    'launch':  ['Launch Date'],
    'aum':     ['AUM (Cr)'],
    'ret1':    ['Return 1Y (%)'],
    'ret3':    ['Return 3Y (%)'],
    'ret5':    ['Return 5Y (%)'],
    'ret10':   ['Return 10Y (%)'],
    'sharpe1': ['Sharpe 1Y'],
    'sortino1':['Sortino 1Y'],
    'alpha1':  ['Alpha 1Y (%)'],
    'info1':   ['Info Ratio 1Y'],
    'sharpe3': ['Sharpe 3Y'],
    'sortino3':['Sortino 3Y'],
    'alpha3':  ['Alpha 3Y (%)'],
    'info3':   ['Info Ratio 3Y'],
    'sharpe5': ['Sharpe 5Y'],
    'sortino5':['Sortino 5Y'],
    'alpha5':  ['Alpha 5Y (%)'],
    'info5':   ['Info Ratio 5Y'],
    'roll1':   ['Median Rolling 1Y (%)'],
    'roll3':   ['Median Rolling 3Y (%)'],
    'roll5':   ['Median Rolling 5Y (%)'],
}

def _norm(s):
    s = str(s).lower()
    for w in ['fund','regular','growth','plan','reg','gr','direct','-','(',')','.',]:
        s = s.replace(w, ' ')
    return ' '.join(s.split())

def safe_float(val):
    try:
        if pd.isna(val): return None
        v = float(val)
        return None if (np.isnan(v) or np.isinf(v)) else v
    except: return None

def fmt(val, suffix='%', dec=2):
    v = safe_float(val)
    if v is None: return '—'
    return f"{v:.{dec}f}{suffix}"

def best_match(name, candidates, threshold=0.52):
    n = _norm(name)
    best, bscore = None, 0
    for c in candidates:
        s = SequenceMatcher(None, n, _norm(str(c))).ratio()
        if s > bscore:
            bscore, best = s, c
    return (best, bscore) if bscore >= threshold else (None, bscore)

def gcol(row, key):
    if row is None: return None
    for k in COLS.get(key, [key]):
        if k in row.index:
            return row[k]
    return None

# ── PDF Styles ─────────────────────────────────────────────────
def S(name='body', size=8, color=None, bold=False, align='LEFT'):
    c = color or colors.black
    return ParagraphStyle(name,
        fontName='Helvetica-Bold' if bold else 'Helvetica',
        fontSize=size, textColor=c,
        leading=size + 2)

def make_pie(labels, values, title='', size=130):
    total = sum(v for v in values if v)
    if not total: return Spacer(1,1)
    margin = 65
    d = Drawing(size + margin*2, size + 35)
    pie = Pie()
    pie.x = margin
    pie.y = 20
    pie.width = pie.height = size * 0.6
    pie.data = [v for v in values if v]
    clean_labels = [l for l, v in zip(labels, values) if v]
    # Show full label for all slices - tiny slices get shorter label
    pie.labels = [
        f"{l}\n{v/total*100:.0f}%"
        if v/total >= 0.05 else f"{l}\n{v/total*100:.0f}%"
        for l, v in zip(clean_labels, pie.data)
    ]
    pie.sideLabels = True
    pie.sideLabelsOffset = 0.25
    pie.simpleLabels = False
    pie.slices.strokeWidth = 0.8
    pie.slices.strokeColor = WHITE
    pie.slices.fontName = 'Helvetica'
    pie.slices.fontSize = 6.5
    for i, col in enumerate(PIE_COLS[:len(pie.data)]):
        pie.slices[i].fillColor = col
    d.add(pie)
    if title:
        d.add(String(size//2 + margin, size + 28, title,
                     fontName='Helvetica-Bold', fontSize=8,
                     fillColor=NAVY, textAnchor='middle'))
    return d

def tbl_style_base():
    return [
        ('BACKGROUND',    (0,0), (-1,0), NAVY),
        ('TEXTCOLOR',     (0,0), (-1,0), WHITE),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 7),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [WHITE, LGREY]),
        ('GRID',          (0,0), (-1,-1), 0.2, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]

# ── Parse client Excel ─────────────────────────────────────────
def parse_client_excel(path):
    xl = pd.ExcelFile(path)
    client_name = 'Client'
    holdings = []

    # Client name
    for sname in xl.sheet_names[:5]:
        df = xl.parse(sname, header=None)
        for i in range(min(6, len(df))):
            row_vals = [str(x) for x in df.iloc[i].tolist()]
            row_str = ' '.join(row_vals)
            m = re.search(r'(?:report|review)\s*[-–]\s*(.+)', row_str, re.I)
            if m:
                client_name = m.group(1).strip(); break
            for val in row_vals[1:3]:
                val = val.strip()
                if (val and val.lower() not in ('nan','','valumonk investment services')
                        and not val.replace('.','').replace('-','').isdigit()
                        and len(val) > 5 and 'Valumonk' not in val):
                    client_name = val; break
            if client_name != 'Client': break
        if client_name != 'Client': break

    # ── Holdings: scan every sheet, pick the one with most valid fund rows ──
    # Keywords used to detect each column type (all matched case-insensitively)
    FUND_KW  = ['fund', 'scheme', 'folio', 'plan']
    VAL_KW   = ['market value', 'current value', 'mkt value', 'curr value',
                'value (rs', 'value(rs', 'mkt val', 'amount', 'corpus']
    CAT_KW   = ['category', 'sub category', 'sub-category', 'asset class', 'type']
    SKIP_KW  = ['total', 'grand total', 'sub total', 'subtotal']

    def _col_match(col_str, keywords):
        """Return True if col_str contains any of the keywords (case-insensitive)."""
        c = col_str.lower().strip()
        return any(kw in c for kw in keywords)

    def _find_header_row(df):
        """Scan rows 0-30 for the row that looks like a column-header row.
        Scores each row by how many cells match fund/value/category keywords."""
        best_row, best_score = -1, 0
        for i in range(min(30, len(df))):
            row_vals = [str(x).strip() for x in df.iloc[i].tolist()]
            score = 0
            for v in row_vals:
                if _col_match(v, FUND_KW):  score += 3   # fund column is essential
                if _col_match(v, VAL_KW):   score += 2
                if _col_match(v, CAT_KW):   score += 1
            if score > best_score:
                best_score, best_row = score, i
        return (best_row, best_score)

    def _extract_holdings(df, hdr_row):
        """Given a df and the index of the header row, extract holdings list."""
        data = df.iloc[hdr_row + 1:].copy()
        # Deduplicate column names by appending _1, _2 ... to dupes
        raw_cols = [str(c).strip() for c in df.iloc[hdr_row].tolist()]
        seen_cols = {}
        deduped = []
        for c in raw_cols:
            if c in seen_cols:
                seen_cols[c] += 1
                deduped.append(f'{c}_{seen_cols[c]}')
            else:
                seen_cols[c] = 0
                deduped.append(c)
        data.columns = deduped
        # Drop blank/unnamed columns
        data = data[[c for c in data.columns
                     if not c.startswith('nan') and c not in ('NaN','','None')]]
        data = data.dropna(how='all')

        # Find which column is fund name / value / category (use original name before dedup suffix)
        def _pick_col(keywords):
            # prefer exact match (no _N suffix), then first suffixed match
            for c in data.columns:
                base = c.split('_')[0] if re.match(r'.+_\d+$', c) else c
                if _col_match(base, keywords) and not re.match(r'.+_\d+$', c):
                    return c
            for c in data.columns:
                base = re.sub(r'_\d+$', '', c)
                if _col_match(base, keywords):
                    return c
            return None

        fund_col = _pick_col(FUND_KW)
        val_col  = _pick_col(VAL_KW)
        cat_col  = _pick_col(CAT_KW)

        if fund_col is None or val_col is None:
            return []

        def _scalar(v):
            """Safely extract a scalar from a value that might be a Series."""
            if isinstance(v, pd.Series):
                v = v.iloc[0] if len(v) else None
            return v

        result = []
        for _, r in data.iterrows():
            fund = str(_scalar(r.get(fund_col, ''))).strip()
            if not fund or fund.lower() in ('nan','none','') \
                    or any(sk in fund.lower() for sk in SKIP_KW):
                continue
            cat = str(_scalar(r.get(cat_col, ''))).strip() if cat_col else ''
            val = _scalar(r.get(val_col, 0))
            try:   val = float(str(val).replace(',','').replace('₹','').replace('Rs','').strip())
            except: val = 0
            if val > 0:
                result.append({
                    'fund': fund,
                    'category': cat if cat.lower() not in ('nan','none','') else 'Other',
                    'market_value': val,
                })
        return result

    # Score every sheet and keep the best
    best_holdings, best_score = [], 0
    for sname in xl.sheet_names:
        try:
            df = xl.parse(sname, header=None)
            hdr_row, score = _find_header_row(df)
            if score < 3:   # needs at least a recognisable fund column
                continue
            candidate = _extract_holdings(df, hdr_row)
            # Weight by both header confidence and number of valid rows found
            weighted = score + len(candidate)
            if weighted > best_score and len(candidate) > 0:
                best_score, best_holdings = weighted, candidate
        except Exception:
            continue

    holdings = best_holdings
    return client_name, holdings

# ── Load analytics ─────────────────────────────────────────────
CATEGORY_FOLDERS = {
    'large cap':'large cap universe','large-cap':'large cap universe','equity: large cap':'large cap universe',
    'mid cap':'mid cap universe','mid-cap':'mid cap universe','equity: mid cap':'mid cap universe',
    'small cap':'small cap universe','small-cap':'small cap universe','equity: small cap':'small cap universe',
    'large & mid':'large and mid cap universe','large and mid':'large and mid cap universe','equity: large and mid cap':'large and mid cap universe','equity: large & mid cap':'large and mid cap universe',
    'flexi cap':'flexi universe','flexi-cap':'flexi universe','equity: flexi cap':'flexi universe',
    'multi cap':'multicap universe','multi-cap':'multicap universe','multicap':'multicap universe',
    'contra':'contra universe','value':'value universe','elss':'elss universe',
    'dividend yield':'dividend yield universe','focused':'focused universe',
    'thematic':'diversified universe','diversified':'diversified universe',
    'aggressive hybrid':'aggressive hybrid universe',
    'balanced advantage':'balanced advantage universe',
    'multi asset':'multi asset universe',
    'conservative hybrid':'conservative hybrid universe',
    'arbitrage':'arbitrage universe','equity savings':'equity savings universe',
}

def load_analytics(analytics_dir):
    dfs, seen = [], set()
    for folder in set(CATEGORY_FOLDERS.values()):
        fpath = os.path.join(analytics_dir, folder)
        if not os.path.isdir(fpath) or fpath in seen: continue
        seen.add(fpath)
        files = sorted([f for f in os.listdir(fpath)
                        if f.endswith('.xlsx') and 'Analytics' in f],
                       key=lambda f: os.path.getmtime(os.path.join(fpath, f)), reverse=True)
        if not files: continue
        try:
            df = pd.read_excel(os.path.join(fpath, files[0]), header=2)
            df['_folder'] = folder
            dfs.append(df.dropna(how='all'))
        except: pass
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

def load_ratings_from_csvs(analytics_dir):
    """Load ratings from per-category CSVs in universe subfolders.
    Returns {fund_name_lower: {rank, score, total, cat_size, category, all_ratings}}
    """
    # CSV locations: each category folder has a Ratings CSV
    CSV_MAP = {
        'Large Cap':         ('large cap universe',          'LargeCap_Ratings.csv'),
        'Mid Cap':           ('mid cap universe',            'MidCap_Ratings.csv'),
        'Small Cap':         ('small cap universe',          'SmallCap_Ratings.csv'),
        'Large & Mid Cap':   ('large and mid cap universe',  'LargeMidCap_Ratings.csv'),
        'Flexi Cap':         ('flexi universe',              'FlexiCap_Ratings.csv'),
        'Multicap':          ('multicap universe',           'Multicap_Ratings.csv'),
        'Contra':            ('contra universe',             'Contra_Ratings.csv'),
        'Value':             ('value universe',              'Value_Ratings.csv'),
        'ELSS':              ('elss universe',               'ELSS_Ratings.csv'),
        'Dividend Yield':    ('dividend yield universe',     'DividendYield_Ratings.csv'),
        'Focused':           ('focused universe',            'Focused_Ratings.csv'),
        'Diversified Funds': ('diversified universe',        'DiversifiedFunds_Ratings.csv'),
        'Aggressive Hybrid': ('aggressive hybrid universe',  'AggressiveHybrid_Ratings.csv'),
        'Balanced Advantage':('balanced advantage universe', 'BalancedAdvantage_Ratings.csv'),
        'Multi Asset':       ('multi asset universe',        'MultiAsset_Ratings.csv'),
        'Conservative Hybrid':('conservative hybrid universe','ConservativeHybrid_Ratings.csv'),
        'Arbitrage':         ('arbitrage universe',          'Arbitrage_Ratings.csv'),
        'Equity Savings':    ('equity savings universe',     'EquitySavings_Ratings.csv'),
    }
    # Categories whose funds ALSO appear in Diversified Funds for dual rating
    DUAL_CATS = {'Flexi Cap','Multicap','Value','ELSS','Contra','Dividend Yield','Focused'}

    # cat_ratings: {category_label: {fund_name_lower: {rank,score,total}}}
    cat_ratings = {}
    for cat_label, (folder, csv_name) in CSV_MAP.items():
        csv_path = os.path.join(analytics_dir, folder, csv_name)
        if not os.path.exists(csv_path):
            continue
        try:
            df = pd.read_csv(csv_path)
            nc = next((c for c in df.columns if 'scheme' in str(c).lower()
                       or 'name' in str(c).lower()), None)
            if nc is None: continue
            entries = {}
            for _, r in df.iterrows():
                name = str(r.get(nc,'')).strip()
                if not name or name.lower() == 'nan': continue
                rank  = safe_float(r.get('Rank'))
                score = safe_float(r.get('Score'))
                total = safe_float(r.get('Total', len(df)))
                entries[name.lower()] = {
                    'rank': rank, 'score': score,
                    'total': int(total) if total else len(df),
                    'cat_size': int(total) if total else len(df),
                    'category': cat_label,
                }
            cat_ratings[cat_label] = entries
        except Exception as e:
            print(f"  CSV load error {csv_path}: {e}")

    # Build flat lookup: fund -> primary entry + all_ratings
    flat = {}
    for cat_label, entries in cat_ratings.items():
        for name_lower, entry in entries.items():
            if name_lower not in flat:
                flat[name_lower] = {**entry, 'all_ratings': []}
            flat[name_lower]['all_ratings'].append({**entry, 'category': cat_label})

    flat['__cat_ratings__'] = cat_ratings
    flat['__dual_cats__'] = DUAL_CATS
    return flat


def get_quartile(rank, cat_size):
    """Return Q1/Q2/Q3/Q4 string based on rank within category."""
    if rank is None or cat_size is None: return '—'
    try:
        r = float(rank); n = float(cat_size)
        if n <= 0: return '—'
        pct = r / n  # rank 1 = top = Q1
        if pct <= 0.25:   return 'Q1'
        elif pct <= 0.50: return 'Q2'
        elif pct <= 0.75: return 'Q3'
        else:             return 'Q4'
    except: return '—'

def quartile_color(q):
    return {
        'Q1': DGREEN,
        'Q2': colors.HexColor('#16a34a'),
        'Q3': AMBER,
        'Q4': RED,
    }.get(q, MGREY)

def match_row(fund_name, analytics_df):
    if analytics_df.empty: return None
    nc = next((c for c in analytics_df.columns
               if 'scheme' in str(c).lower() or 'name' in str(c).lower()), None)
    if nc is None: return None
    cands = analytics_df[nc].astype(str).tolist()
    match, score = best_match(fund_name, cands)
    if match and score > 0.52:
        rows = analytics_df[analytics_df[nc] == match]
        return rows.iloc[0] if not rows.empty else None
    return None


# Map client Excel category names -> our ratings category label
CLIENT_CAT_MAP = {
    'large cap': 'Large Cap', 'large-cap': 'Large Cap', 'largecap': 'Large Cap',
    'mid cap': 'Mid Cap', 'mid-cap': 'Mid Cap', 'midcap': 'Mid Cap',
    'small cap': 'Small Cap', 'small-cap': 'Small Cap', 'smallcap': 'Small Cap',
    'large & mid': 'Large & Mid Cap', 'large and mid cap': 'Large & Mid Cap',
    'flexi cap': 'Flexi Cap', 'flexi-cap': 'Flexi Cap', 'flexicap': 'Flexi Cap',
    'multi cap': 'Multicap', 'multi-cap': 'Multicap', 'multicap': 'Multicap',
    'contra': 'Contra', 'value': 'Value', 'elss': 'ELSS',
    'dividend yield': 'Dividend Yield', 'focused': 'Focused',
    'thematic': 'Diversified Funds', 'diversified': 'Diversified Funds',
    'aggressive hybrid': 'Aggressive Hybrid',
    'balanced advantage': 'Balanced Advantage',
    'multi asset': 'Multi Asset',
    'conservative hybrid': 'Conservative Hybrid',
    'arbitrage': 'Arbitrage', 'equity savings': 'Equity Savings',
}

def enrich(holdings, analytics_df, ratings):
    enriched = []
    for h in holdings:
        row = match_row(h['fund'], analytics_df)
        def g(k): return safe_float(gcol(row, k))

        launch_raw = gcol(row, 'launch')
        try: launch = pd.to_datetime(launch_raw).strftime('%b %Y') if launch_raw else '—'
        except: launch = '—'

        ri, ra = {}, []
        cat_ratings = ratings.get('__cat_ratings__', {})
        dual_cats   = ratings.get('__dual_cats__', set())

        # Map client category -> our ratings category label
        client_cat_norm = h['category'].lower().strip()
        cat_label = CLIENT_CAT_MAP.get(client_cat_norm)

        # If no direct map, fuzzy match
        if not cat_label:
            best_cat, best_s = None, 0
            for cl in cat_ratings:
                s = SequenceMatcher(None, client_cat_norm, cl.lower()).ratio()
                if s > best_s: best_s, best_cat = s, cl
            cat_label = best_cat if best_s > 0.5 else None

        # Fuzzy match fund name ONLY within the correct category
        if cat_label and cat_label in cat_ratings:
            entries = cat_ratings[cat_label]
            mk, ms = best_match(h['fund'], list(entries.keys()), 0.45)
            if mk:
                ri = entries[mk]

        # Build all_ratings: own category + Diversified Funds if eligible
        if ri:
            ra = [ri]
            if cat_label in dual_cats and 'Diversified Funds' in cat_ratings:
                div_entries = cat_ratings['Diversified Funds']
                mk2, ms2 = best_match(h['fund'], list(div_entries.keys()), 0.45)
                if mk2 and ms2 > 0.45:
                    div_ri = div_entries[mk2]
                    if div_ri != ri:
                        ra.append(div_ri)

        # Quartile and score from the matched category entry
        rank     = ri.get('rank')
        total    = ri.get('total') or ri.get('cat_size')
        score    = safe_float(ri.get('score'))
        quartile = get_quartile(rank, total)

        # All-ratings quartiles
        ra_with_q = []
        for r in ra:
            q = get_quartile(r.get('rank'), r.get('cat_size'))
            ra_with_q.append({**r, 'quartile': q})

        v_lbl, v_col = ('N/A', MGREY)
        r1, m1 = g('ret1'), g('roll1')
        if r1 is not None and m1 is not None:
            diff = r1 - m1
            v_lbl = 'Outperforming' if diff >= 2 else ('Underperforming' if diff <= -2 else 'Inline')
            v_col = DGREEN if diff >= 2 else (RED if diff <= -2 else AMBER)

        # Assessment follows quartile — consistent with ranking
        if not ri:  # unrated
            cap_lbl, cap_col = 'Unrated', MGREY
        elif quartile == 'Q1': cap_lbl, cap_col = 'Well Deployed', DGREEN
        elif quartile == 'Q2': cap_lbl, cap_col = 'Acceptable', AMBER
        elif quartile == 'Q3': cap_lbl, cap_col = 'Review Needed', RED
        elif quartile == 'Q4': cap_lbl, cap_col = 'Review Needed', RED
        else: cap_lbl, cap_col = 'Unrated', MGREY

        enriched.append({**h,
            'ret1':g('ret1'),'ret3':g('ret3'),'ret5':g('ret5'),'ret10':g('ret10'),
            'sharpe1':g('sharpe1'),'sortino1':g('sortino1'),'alpha1':g('alpha1'),'info1':g('info1'),
            'sharpe3':g('sharpe3'),'sortino3':g('sortino3'),'alpha3':g('alpha3'),'info3':g('info3'),
            'sharpe5':g('sharpe5'),'sortino5':g('sortino5'),'alpha5':g('alpha5'),'info5':g('info5'),
            'roll1':g('roll1'),'roll3':g('roll3'),'roll5':g('roll5'),
            'aum':g('aum'), 'launch':launch,
            'quartile':quartile,
            'quartile_col': quartile_color(quartile),
            'all_ratings':ra_with_q,
            'score':score,
            'cap_verdict':cap_lbl,'cap_col':cap_col,
        })
    return enriched

# ── GENERATE PDF ───────────────────────────────────────────────
def generate_pdf(client_name, enriched, cat_breakdown, total_mv, date_str, output_dir):
    out = os.path.join(output_dir, f"Portfolio_Review_{client_name.replace(' ','_')}_{date_str.replace(' ','')}.pdf")
    PAGE = landscape(A4)
    W = PAGE[0] - 24*mm

    doc = SimpleDocTemplate(out, pagesize=PAGE,
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=10*mm, bottomMargin=10*mm)
    story = []

    # ── PAGE 1: Summary + Allocation ──────────────────────────
    # Logo + company name header
    from reportlab.platypus import Image as RLImage
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'valumonk_logo.jpg')
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=28*mm, height=22*mm)
    else:
        logo = Spacer(28*mm, 22*mm)

    header_tbl = Table([[
        logo,
        Spacer(1,1),
        Paragraph(f'Report Date: {date_str}', S('sm', 8, MGREY)),
    ]], colWidths=[32*mm, 80*mm, W-112*mm])
    header_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (2,0), (2,0), 'RIGHT'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 5))
    story.append(HRFlowable(width='100%', thickness=1, color=GREEN, spaceAfter=6))
    story.append(Paragraph(f'Portfolio Review — {client_name}',
                            S('title', 20, NAVY, bold=True)))
    story.append(HRFlowable(width='100%', thickness=1.5, color=NAVY, spaceAfter=8))

    # KPIs
    kpi_data = [[
        [Paragraph('Total Portfolio Value', S('kl',7,MGREY)),
         Spacer(1,2),
         Paragraph(f"Rs. {total_mv:,.0f}", S('kv',14,NAVY,True))],
        [Paragraph('No. of Holdings', S('kl',7,MGREY)),
         Spacer(1,2),
         Paragraph(str(len(enriched)), S('kv',14,NAVY,True))],
        [Paragraph('Categories', S('kl',7,MGREY)),
         Spacer(1,2),
         Paragraph(str(len(cat_breakdown)), S('kv',14,NAVY,True))],
        [Paragraph('Report Date', S('kl',7,MGREY)),
         Spacer(1,2),
         Paragraph(date_str, S('kv',14,NAVY,True))],
    ]]
    kt = Table(kpi_data, colWidths=[W/4]*4)
    kt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#EEF2FF')),
        ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#c7d2fe')),
        ('INNERGRID',(0,0),(-1,-1),0.5,colors.HexColor('#c7d2fe')),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LEFTPADDING',(0,0),(-1,-1),12),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(kt)
    story.append(Spacer(1, 10))

    # Allocation: pie + table side by side
    story.append(Paragraph('Portfolio Allocation by Category', S('h2',11,NAVY,True)))
    story.append(Spacer(1, 4))

    pie_labels = [c for c,d in sorted(cat_breakdown.items(), key=lambda x:-x[1]['value'])]
    pie_values = [cat_breakdown[c]['value'] for c in pie_labels]
    pie_chart  = make_pie(pie_labels, pie_values, size=150)

    cat_hdr = [[Paragraph(x, S('lh',7,WHITE,True)) for x in
                ['Category','Market Value (Rs.)','Mix %','No. of Funds']]]
    cat_rows = cat_hdr[:]
    for cat in pie_labels:
        d = cat_breakdown[cat]
        pct = d['value']/total_mv*100 if total_mv else 0
        cat_rows.append([
            Paragraph(cat, S('b',8)),
            Paragraph(f"Rs. {d['value']:,.0f}", S('b',8)),
            Paragraph(f"{pct:.1f}%", S('b',8)),
            Paragraph(str(len(d['funds'])), S('b',8)),
        ])
    cat_rows.append([
        Paragraph('TOTAL', S('bold',8,NAVY,True)),
        Paragraph(f"Rs. {total_mv:,.0f}", S('bold',8,NAVY,True)),
        Paragraph('100.0%', S('bold',8,NAVY,True)),
        Paragraph(str(len(enriched)), S('bold',8,NAVY,True)),
    ])
    ct = Table(cat_rows, colWidths=[55*mm, 45*mm, 25*mm, 25*mm])
    ct.setStyle(TableStyle([
        *tbl_style_base(),
        ('BACKGROUND',(0,-1),(-1,-1), colors.HexColor('#dbeafe')),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
    ]))

    side = Table([[pie_chart, ct]], colWidths=[100*mm, W-100*mm])
    side.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                               ('ALIGN',(0,0),(0,-1),'CENTRE'),
                               ('LEFTPADDING',(0,0),(0,-1),0),
                               ('RIGHTPADDING',(0,0),(0,-1),8)]))
    story.append(side)
    story.append(PageBreak())

    # ── PAGE 2: Performance by Category ───────────────────────
    story.append(Paragraph('Performance Overview — by Category', S('h2',12,NAVY,True)))
    story.append(Spacer(1, 6))

    perf_hdr = [Paragraph(x, S('lh',7,WHITE,True)) for x in
                ['Fund','Quartile','All Category Quartiles',
                 'Market Value','1Y','3Y','5Y','Cat Med 1Y']]

    for cat, d in sorted(cat_breakdown.items(), key=lambda x:-x[1]['value']):
        cat_title = Table([[Paragraph(f"  {cat}", S('ct',8,WHITE,True))]],
                          colWidths=[W])
        cat_title.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1), NAVY),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('LEFTPADDING',(0,0),(-1,-1),6),
        ]))
        rows = [perf_hdr]
        for e in sorted(d['funds'], key=lambda x:-x['market_value']):
            ra = e.get('all_ratings',[])
            if len(ra) > 1:
                r_str = ' | '.join([f"{r['quartile']} in {r['category']}"
                                    for r in ra if r.get('quartile','—') != '—'])
            else:
                r_str = e['quartile']
            qc = e['quartile_col']
            rows.append([
                Paragraph(e['fund'], S('b',7)),
                Paragraph(e['quartile'], S('v',8,qc,True)),
                Paragraph(r_str or '—', S('sm',6,MGREY)),
                Paragraph(f"Rs. {e['market_value']:,.0f}", S('b',7)),
                Paragraph(fmt(e['ret1']), S('b',7)),
                Paragraph(fmt(e['ret3']), S('b',7)),
                Paragraph(fmt(e['ret5']), S('b',7)),
                Paragraph(fmt(e['roll1']), S('b',7,MGREY)),
            ])
        tbl = Table(rows, colWidths=[65*mm,15*mm,55*mm,28*mm,15*mm,15*mm,15*mm,20*mm])
        tbl.setStyle(TableStyle(tbl_style_base()))
        story.append(KeepTogether([cat_title, Spacer(1,1), tbl, Spacer(1,6)]))

    story.append(PageBreak())

    # ── PAGE 3: Capital Deployment ─────────────────────────────
    # Mini header
    if os.path.exists(logo_path):
        mini_logo = RLImage(logo_path, width=10*mm, height=8*mm)
    else:
        mini_logo = Spacer(10*mm, 8*mm)
    mini_hdr = Table([[mini_logo,
                       Paragraph('ValuMonk Investment Services', S('vm',9,NAVY,True)),
                       Paragraph(f'{client_name}  ·  {date_str}', S('sm',7,MGREY))]],
                     colWidths=[14*mm, 80*mm, W-94*mm])
    mini_hdr.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',(2,0),(2,0),'RIGHT'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(mini_hdr)
    story.append(HRFlowable(width='100%',thickness=0.5,color=GREEN,spaceAfter=4))

    well  = sum(e['market_value'] for e in enriched if e['cap_verdict']=='Well Deployed')
    acc   = sum(e['market_value'] for e in enriched if e['cap_verdict']=='Acceptable')
    rev   = sum(e['market_value'] for e in enriched if e['cap_verdict']=='Review Needed')
    unr   = sum(e['market_value'] for e in enriched if e['cap_verdict']=='Unrated')

    story.append(Paragraph('Capital Deployment Analysis', S('h2',12,NAVY,True)))
    story.append(Paragraph(
        'Two independent measures of fund quality are shown below:',
        S('sm',8,NAVY,True)))
    story.append(Spacer(1,3))
    story.append(Paragraph(
        'Quartile (Q1-Q4): relative ranking within the fund\'s own category. '
        'Q1 = top 25% of peers, Q4 = bottom 25%. A fund can be Q1 in a weak category '
        'while still having a low absolute score — meaning it is the best of its peers '
        'but the category overall may need review.',
        S('sm',7,MGREY)))
    story.append(Spacer(1,2))
    story.append(Paragraph(
        'Score /100: absolute Valumonk quantitative score based on rolling returns, '
        'alpha, sortino ratio, info ratio and capture ratios across 1Y/3Y/5Y periods. '
        'Higher is better. Use this to assess absolute fund quality independent of peers.',
        S('sm',7,MGREY)))
    story.append(Spacer(1,6))

    # Summary pie + summary table
    sum_labels = [l for l,v in [('Well Deployed',well),('Acceptable',acc),
                                  ('Review Needed',rev),('Unrated',unr)] if v>0]
    sum_values = [v for v in [well,acc,rev,unr] if v>0]
    sum_pie = make_pie(sum_labels, sum_values, 'Capital by Quality', size=90)

    sum_hdr = [[Paragraph(x, S('lh',7,WHITE,True)) for x in
                ['Assessment','Market Value (Rs.)','% of Portfolio']]]
    sum_rows = sum_hdr[:]
    for lbl, val, col in [('Well Deployed',well,DGREEN),('Acceptable',acc,AMBER),
                           ('Review Needed',rev,RED),('Unrated',unr,MGREY)]:
        if val <= 0: continue
        sum_rows.append([
            Paragraph(lbl, S('sl',8,col,True)),
            Paragraph(f"Rs. {val:,.0f}", S('b',8)),
            Paragraph(f"{val/total_mv*100:.1f}%" if total_mv else '—', S('b',8)),
        ])
    st = Table(sum_rows, colWidths=[50*mm,45*mm,35*mm])
    st.setStyle(TableStyle(tbl_style_base()))

    sum_side = Table([[sum_pie, st]], colWidths=[120*mm, W-120*mm])
    sum_side.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                                   ('ALIGN',(0,0),(0,-1),'CENTRE'),
                                   ('LEFTPADDING',(0,0),(0,-1),0),
                                   ('RIGHTPADDING',(0,0),(0,-1),8)]))
    story.append(sum_side)
    story.append(Spacer(1, 8))

    # Fund-level capital table
    cap_hdr = [Paragraph(x, S('lh',7,WHITE,True)) for x in
               ['Fund','Category','Market Value','Launch','Quartile','Score /100','Assessment']]
    cap_rows = [cap_hdr]
    for e in sorted(enriched, key=lambda x:-x['market_value']):
        score = e['score']
        qc    = e['quartile_col']
        if score and safe_float(score):
            bar_w = 70
            pct = max(0, min(1, float(score)/100))
            col = DGREEN if pct>=0.7 else (AMBER if pct>=0.5 else RED)
            d = Drawing(bar_w+40, 12)
            d.add(Rect(0,1,bar_w,10,fillColor=LGREY,strokeColor=None))
            d.add(Rect(0,1,bar_w*pct,10,fillColor=col,strokeColor=None))
            d.add(String(bar_w+2,3,f"{score:.0f}/100",
                         fontName='Helvetica-Bold',fontSize=7,fillColor=NAVY))
            score_cell = d
        else:
            score_cell = Paragraph('—', S('b',7))
        cap_rows.append([
            Paragraph(e['fund'], S('b',7)),
            Paragraph(e['category'], S('sm',6,MGREY)),
            Paragraph(f"Rs. {e['market_value']:,.0f}", S('b',7)),
            Paragraph(e['launch'], S('sm',6,MGREY)),
            Paragraph(e['quartile'], S('cv',8,qc,True)),
            score_cell,
            Paragraph(e['cap_verdict'], S('cv',7,e['cap_col'],True)),
        ])
    cap_tbl = Table(cap_rows, colWidths=[65*mm,22*mm,28*mm,18*mm,15*mm,42*mm,30*mm])
    cap_tbl.setStyle(TableStyle(tbl_style_base()))
    story.append(cap_tbl)
    story.append(PageBreak())

    # ── PAGE 4: Detailed Metrics ───────────────────────────────
    # Mini header
    if os.path.exists(logo_path):
        mini_logo = RLImage(logo_path, width=10*mm, height=8*mm)
    else:
        mini_logo = Spacer(10*mm, 8*mm)
    mini_hdr = Table([[mini_logo,
                       Paragraph('ValuMonk Investment Services', S('vm',9,NAVY,True)),
                       Paragraph(f'{client_name}  ·  {date_str}', S('sm',7,MGREY))]],
                     colWidths=[14*mm, 80*mm, W-94*mm])
    mini_hdr.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',(2,0),(2,0),'RIGHT'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(mini_hdr)
    story.append(HRFlowable(width='100%',thickness=0.5,color=GREEN,spaceAfter=4))

    story.append(Paragraph('Detailed Fund Metrics', S('h2',12,NAVY,True)))
    story.append(Paragraph(
        'Sharpe, Sortino, Alpha and Info Ratio for 1Y / 3Y / 5Y periods. '
        'All from latest Valumonk analytics run.',
        S('sm',7,MGREY)))
    story.append(Spacer(1, 6))

    det_hdr = [Paragraph(x, S('lh',7,WHITE,True)) for x in
               ['Fund','Category','1Y Ret','3Y Ret','5Y Ret',
                'Sharpe 1Y','Sortino 1Y','Alpha 1Y','Info R. 1Y',
                'Sharpe 3Y','Sortino 3Y','Alpha 3Y',
                'Sharpe 5Y','Sortino 5Y','Alpha 5Y']]
    det_rows = [det_hdr]
    for e in sorted(enriched, key=lambda x:(x['category'],-x['market_value'])):
        det_rows.append([
            Paragraph(e['fund'], S('b',6.5)),
            Paragraph(e['category'], S('sm',6,MGREY)),
            Paragraph(fmt(e['ret1']), S('b',7)),
            Paragraph(fmt(e['ret3']), S('b',7)),
            Paragraph(fmt(e['ret5']), S('b',7)),
            Paragraph(fmt(e['sharpe1'],'',2), S('b',7)),
            Paragraph(fmt(e['sortino1'],'',2), S('b',7)),
            Paragraph(fmt(e['alpha1']), S('b',7)),
            Paragraph(fmt(e['info1'],'',2), S('b',7)),
            Paragraph(fmt(e['sharpe3'],'',2), S('b',7)),
            Paragraph(fmt(e['sortino3'],'',2), S('b',7)),
            Paragraph(fmt(e['alpha3']), S('b',7)),
            Paragraph(fmt(e['sharpe5'],'',2), S('b',7)),
            Paragraph(fmt(e['sortino5'],'',2), S('b',7)),
            Paragraph(fmt(e['alpha5']), S('b',7)),
        ])

    cw_name = 60*mm
    cw_cat  = 20*mm
    cw_rest = (W - cw_name - cw_cat) / 13
    det_tbl = Table(det_rows, colWidths=[cw_name, cw_cat] + [cw_rest]*13)
    det_tbl.setStyle(TableStyle(tbl_style_base()))
    story.append(det_tbl)

    story.append(Spacer(1,10))
    story.append(HRFlowable(width='100%',thickness=0.5,color=MGREY))
    story.append(Spacer(1,4))
    story.append(Paragraph(
        'This report is prepared by Valumonk Investment Services for advisor use only. '
        'Mutual fund investments are subject to market risk. '
        'Past performance is not indicative of future returns. '
        'Regular Growth plans only. Ratings based on Valumonk quantitative framework.',
        S('disc',7,MGREY)))

    doc.build(story)
    return out

# ── GENERATE EXCEL ─────────────────────────────────────────────
def generate_excel(client_name, enriched, cat_breakdown, total_mv, date_str, output_dir):
    out = os.path.join(output_dir,
          f"Portfolio_Review_{client_name.replace(' ','_')}_{date_str.replace(' ','')}.xlsx")
    wb = openpyxl.Workbook()

    # Excel styles
    hdr_fill = PatternFill('solid', fgColor='1E2B4A')
    grn_fill = PatternFill('solid', fgColor='2DB84B')
    lgr_fill = PatternFill('solid', fgColor='F1F5F9')
    hdr_font = Font(bold=True, color='FFFFFF', size=9)
    ttl_font = Font(bold=True, color='1E2B4A', size=10)
    reg_font = Font(size=9)
    ctr = Alignment(horizontal='center', vertical='center')
    lft = Alignment(horizontal='left', vertical='center')
    thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    def hrow(ws, row_idx, values, widths=None):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = ctr; c.border = thin
        if widths:
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    def drow(ws, row_idx, values, alt=False):
        fill = lgr_fill if alt else PatternFill('solid', fgColor='FFFFFF')
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            c.fill = fill; c.font = reg_font
            c.alignment = lft; c.border = thin

    def verdict_color(ws, row_idx, col_idx, verdict):
        fc = {'Outperforming':'16a34a','Inline':'d97706',
              'Underperforming':'dc2626','N/A':'64748b'}.get(verdict,'64748b')
        ws.cell(row=row_idx, column=col_idx).font = Font(bold=True, color=fc, size=9)

    def cap_color(ws, row_idx, col_idx, verdict):
        fc = {'Well Deployed':'16a34a','Acceptable':'d97706',
              'Review Needed':'dc2626','Unrated':'64748b'}.get(verdict,'64748b')
        ws.cell(row=row_idx, column=col_idx).font = Font(bold=True, color=fc, size=9)

    # ── Sheet 1: Summary ──
    ws = wb.active; ws.title = 'Summary'
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Portfolio Review — {client_name}'
    ws['A1'].font = Font(bold=True, color='1E2B4A', size=14)
    ws['A2'] = f'Report Date: {date_str}  |  Regular Growth Plans'
    ws['A2'].font = Font(color='64748b', size=9)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 18

    hrow(ws, 3, ['Metric','Value'], [30, 25])
    metrics = [
        ('Total Portfolio Value', f"Rs. {total_mv:,.0f}"),
        ('No. of Holdings', len(enriched)),
        ('No. of Categories', len(cat_breakdown)),
        ('Report Generated', date_str),
    ]
    for i, (k, v) in enumerate(metrics):
        drow(ws, 4+i, [k, v], i%2==1)

    ws.row_dimensions[9].height = 6

    # Category breakdown
    hrow(ws, 10, ['Category','Market Value (Rs.)','Mix %','No. of Funds'],
         [22, 22, 12, 15])
    for i, (cat, d) in enumerate(sorted(cat_breakdown.items(),
                                         key=lambda x:-x[1]['value'])):
        pct = d['value']/total_mv*100 if total_mv else 0
        drow(ws, 11+i, [cat, d['value'], round(pct,1), len(d['funds'])], i%2==1)
        ws.cell(row=11+i, column=2).number_format = '#,##0'
        ws.cell(row=11+i, column=3).number_format = '0.0"%"'
    tot_row = 11 + len(cat_breakdown)
    drow(ws, tot_row, ['TOTAL', total_mv, 100.0, len(enriched)])
    for ci in range(1,5):
        ws.cell(row=tot_row, column=ci).font = Font(bold=True, color='1E2B4A', size=9)
        ws.cell(row=tot_row, column=ci).fill = PatternFill('solid', fgColor='DBEAFE')
    ws.cell(row=tot_row, column=2).number_format = '#,##0'
    ws.cell(row=tot_row, column=3).number_format = '0.0"%"'

    # ── Sheet 2: Performance ──
    ws2 = wb.create_sheet('Performance')
    hrow(ws2, 1,
         ['Fund','Category','Market Value (Rs.)','Launch','Quartile',
          'All Category Quartiles','1Y Ret %','3Y Ret %','5Y Ret %',
          'Cat Median 1Y %'],
         [40,16,18,12,10,45,10,10,10,14])
    for i, e in enumerate(sorted(enriched, key=lambda x:(x['category'],-x['market_value']))):
        ra = e.get('all_ratings',[])
        r_str = ' | '.join([f"{r['quartile']} in {r['category']}"
                            for r in ra if r.get('quartile','—')!='—']) if len(ra)>1 else e['quartile']
        row_vals = [
            e['fund'], e['category'], e['market_value'], e['launch'], e['quartile'],
            r_str,
            safe_float(e['ret1']), safe_float(e['ret3']), safe_float(e['ret5']),
            safe_float(e['roll1']),
        ]
        drow(ws2, 2+i, row_vals, i%2==1)
        ws2.cell(row=2+i, column=3).number_format = '#,##0'
        for col in [7,8,9,10]:
            ws2.cell(row=2+i, column=col).number_format = '0.00'
        # Colour quartile cell
        qfc = {'Q1':'16a34a','Q2':'2DB84B','Q3':'d97706','Q4':'dc2626'}.get(e['quartile'],'64748b')
        ws2.cell(row=2+i, column=5).font = Font(bold=True, color=qfc, size=9)

    # ── Sheet 3: Capital Deployment ──
    ws3 = wb.create_sheet('Capital Deployment')
    hrow(ws3, 1,
         ['Fund','Category','Market Value (Rs.)','Launch',
          'Quartile','Score /100','Assessment'],
         [40,16,18,12,10,12,16])
    for i, e in enumerate(sorted(enriched, key=lambda x:-x['market_value'])):
        score = safe_float(e['score'])
        drow(ws3, 2+i, [
            e['fund'], e['category'], e['market_value'], e['launch'],
            e['quartile'], score, e['cap_verdict']
        ], i%2==1)
        ws3.cell(row=2+i, column=3).number_format = '#,##0'
        qfc = {'Q1':'16a34a','Q2':'2DB84B','Q3':'d97706','Q4':'dc2626'}.get(e['quartile'],'64748b')
        ws3.cell(row=2+i, column=5).font = Font(bold=True, color=qfc, size=9)
        cap_color(ws3, 2+i, 7, e['cap_verdict'])

    # ── Sheet 4: Detailed Metrics ──
    ws4 = wb.create_sheet('Detailed Metrics')
    hrow(ws4, 1,
         ['Fund','Category','1Y Ret %','3Y Ret %','5Y Ret %',
          'Sharpe 1Y','Sortino 1Y','Alpha 1Y %','Info R. 1Y',
          'Sharpe 3Y','Sortino 3Y','Alpha 3Y %',
          'Sharpe 5Y','Sortino 5Y','Alpha 5Y %'],
         [40,16,10,10,10,10,10,10,10,10,10,10,10,10,10])
    for i, e in enumerate(sorted(enriched, key=lambda x:(x['category'],-x['market_value']))):
        drow(ws4, 2+i, [
            e['fund'], e['category'],
            safe_float(e['ret1']), safe_float(e['ret3']), safe_float(e['ret5']),
            safe_float(e['sharpe1']), safe_float(e['sortino1']),
            safe_float(e['alpha1']), safe_float(e['info1']),
            safe_float(e['sharpe3']), safe_float(e['sortino3']),
            safe_float(e['alpha3']),
            safe_float(e['sharpe5']), safe_float(e['sortino5']),
            safe_float(e['alpha5']),
        ], i%2==1)
        for col in range(3,16):
            ws4.cell(row=2+i, column=col).number_format = '0.00'

    # Freeze panes on all data sheets
    for ws_x in [ws2, ws3, ws4]:
        ws_x.freeze_panes = ws_x.cell(row=2, column=1)

    wb.save(out)
    return out

# ── MAIN: Returns zip with both files ─────────────────────────
def generate_review(client_excel_path, analytics_dir, ratings_xlsx, output_dir):
    client_name, holdings = parse_client_excel(client_excel_path)
    date_str = datetime.today().strftime('%d %b %Y')

    analytics_df = load_analytics(analytics_dir)
    ratings      = load_ratings_from_csvs(analytics_dir)
    enriched     = enrich(holdings, analytics_df, ratings)

    total_mv = sum(e['market_value'] for e in enriched)
    cat_breakdown = {}
    for e in enriched:
        c = e['category']
        if c not in cat_breakdown:
            cat_breakdown[c] = {'value':0,'funds':[]}
        cat_breakdown[c]['value'] += e['market_value']
        cat_breakdown[c]['funds'].append(e)

    pdf_path  = generate_pdf(client_name, enriched, cat_breakdown,
                             total_mv, date_str, output_dir)
    xlsx_path = generate_excel(client_name, enriched, cat_breakdown,
                               total_mv, date_str, output_dir)

    # Zip both
    zip_name = f"Portfolio_Review_{client_name.replace(' ','_')}_{date_str.replace(' ','')}.zip"
    zip_path = os.path.join(output_dir, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(pdf_path,  os.path.basename(pdf_path))
        zf.write(xlsx_path, os.path.basename(xlsx_path))
    os.unlink(pdf_path)
    os.unlink(xlsx_path)
    return zip_path

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Usage: python portfolio_review.py <client.xlsx>"); sys.exit(1)
    base = r'C:\Users\gargi\OneDrive\Desktop\research_project'
    out  = generate_review(sys.argv[1], base,
                           os.path.join(base,'MF_Ratings_All.xlsx'), base)
    print(f"Generated: {out}")
